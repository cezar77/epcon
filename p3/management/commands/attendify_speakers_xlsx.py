# -*- coding: utf-8 -*-
""" Update an Attendify speakers XLSX file with the current list of
    speakers.

    Usage: manage.py attendify_speakers_xlsx ep2016 speakers.xlsx

    Note that for Attendify you have to download the speakers before
    running this script, since they add meta data to the downloaded
    file which has to be kept around when uploading it again.

    The script updates speakers.xlsx in place. Unfortunately, Attendify
    currently has a bug in that it doesn't accept the file format
    generated by openpyxl. Opening the file in LibreOffice and saving
    it (without changes) fixes this as work-around.

    Attendify Worksheet "Schedule" format
    -------------------------------------

    Row A4: First Name, Last Name, Company (Optional), Position
    (Optional), Group (Optional). Profile (Optional), Email
    (Optional), Phone (Optional), Twitter (Optional), Facebook
    (Optional), LinkedIn (Optional), Google+ (Optional), UID (do not
    delete)
    
    Row A6: Start of data

"""
from django.core.management.base import BaseCommand, CommandError
from django.core import urlresolvers
from django.conf import settings
from django.utils.html import strip_tags
from conference import models as cmodels
from conference import utils
from p3 import models

import datetime
from collections import defaultdict
from optparse import make_option
import operator
import markdown2
import openpyxl
import urlparse

### Globals

# Debug output ?
_debug = 1

# Website URL to use for making the profile links absolute
WEBSITE_URL = 'https://ep2018.europython.eu/'

# These must match the talk .type or .admin_type
from accepted_talks import TYPE_NAMES

### Helpers

def profile_url(user):

    return urlparse.urljoin(
        WEBSITE_URL,
        urlresolvers.reverse('conference-profile',
                             args=[user.attendeeprofile.slug]))

def format_text(text, remove_tags=False, output_html=True):

    # Remove whitespace
    text = text.strip()
    if not text:
        return text

    # Remove links, tags, etc.
    if remove_tags:
        text = strip_tags(text)

    # Remove quotes
    if text[0] == '"' and text[-1] == '"':
        text = text[1:-1]

    # Convert markdown markup to HTML
    if output_html:
        text = markdown2.markdown(text)

    return text    

def add_speaker(data, speaker):

    # Get speaker profile
    user = speaker.user
    profile = cmodels.AttendeeProfile.objects.get(user=user)
    p3profile = models.P3Profile.objects.get(profile=profile)

    # Skip speakers without public profile. Speaker profiles must be
    # public, but you never know. See conference/models.py
    if profile.visibility != 'p':
        return

    # Collect data
    first_name = speaker.user.first_name.title()
    last_name = speaker.user.last_name.title()
    company = profile.company
    position = profile.job_title
    profile_text = (u'<a href="%s%s">Profile on EuroPython Website</a>' %
                    (settings.DEFAULT_URL_PREFIX, profile_url(user)))
    twitter = p3profile.twitter
    if twitter.startswith(('https://twitter.com/', 'http://twitter.com/')):
        twitter = twitter.split('/')[-1]

    # Skip special entries
    full_name = first_name + last_name
    if first_name == 'To Be' and last_name == 'Announced':
        return
   
    # UID
    uid = u''
    
    data.append((
        first_name,
        last_name,
        company,
        position,
        u'', # group
        profile_text,
        u'', # email: not published
        u'', # phone: not published
        twitter,
        u'', # facebook
        u'', # linkedin
        u'', # google+
        uid))

# Start row of data in spreadsheet (Python 0-based index)
SPEAKERS_WS_START_DATA = 5

# Column number of UID columns (Python 0-based index)
SPEAKERS_UID_COLUMN = 12

# Number of columns to make row unique (first, last, company)
SPEAKERS_UNIQUE_COLS = 3

def update_speakers(speakers_xlsx, new_data, updated_xlsx=None):

    # Load workbook
    wb = openpyxl.load_workbook(speakers_xlsx)
    assert wb.sheetnames == [u'Instructions', u'Speakers', u'System']
    ws = wb['Speakers']

    # Extract data values
    ws_data = list(ws.values)[SPEAKERS_WS_START_DATA:]
    print ('read %i data lines' % len(ws_data))
    print ('first line: %r' % ws_data[:1])
    print ('last line: %r' % ws_data[-1:])

    # Reconcile UIDs / talks
    uids = {}
    for line in ws_data:
        uid = line[SPEAKERS_UID_COLUMN]
        if not uid:
            continue
        uids[tuple(line[:SPEAKERS_UNIQUE_COLS])] = uid

    # Add UID to new data
    new_speakers = []
    for line in new_data:
        key = tuple(line[:SPEAKERS_UNIQUE_COLS])
        if key not in uids:
            print ('New speaker %s found' % (key,))
            uid = u''
        else:
            uid = uids[key]
        line = tuple(line[:SPEAKERS_UID_COLUMN]) + (uid,)
        new_speakers.append(line)
    new_data = new_speakers

    # Replace old data with new data
    old_data_rows = len(ws_data)
    new_data_rows = len(new_data)
    print ('new data: %i data lines' % new_data_rows)
    offset = SPEAKERS_WS_START_DATA + 1
    print ('new_data = %i rows' % len(new_data))
    for j, row in enumerate(ws[offset: offset + new_data_rows - 1]):
        new_row = new_data[j]
        if _debug:
            print ('updating row %i with %r' % (j, new_row))
        if len(row) > len(new_row):
            row = row[:len(new_row)]
        for i, cell in enumerate(row):
            cell.value = new_row[i]
    
    # Overwrite unused cells with None
    if new_data_rows < old_data_rows:
        for j, row in enumerate(ws[offset + new_data_rows + 1:
                                   offset + old_data_rows + 1]):
            if _debug:
                print ('clearing row %i' % (j,))
            for i, cell in enumerate(row):
                cell.value = None

    # Write updated data
    if updated_xlsx is None:
        updated_xlsx = speakers_xlsx
    wb.save(updated_xlsx)
    
###

class Command(BaseCommand):
    option_list = BaseCommand.option_list + (
        # make_option('--option',
        #     action='store',
        #     dest='option_attr',
        #     default=0,
        #     type='int',
        #     help='Help text',
        # ),
    )

    args = '<conference> <xlsx-file>'

    def handle(self, *args, **options):
        try:
            conference = args[0]
        except IndexError:
            raise CommandError('conference not specified')
        try:
            speakers_xlsx = args[1]
        except IndexError:
            raise CommandError('XLSX file not specified')

        # Get speaker records
        speakers = set()
        talks = cmodels.Talk.objects.accepted(conference)
        for t in talks:
            speakers |= set(t.get_all_speakers())

        # Collect profiles
        data = []
        for speaker in speakers:
            add_speaker(data, speaker)
        data.sort()

        # Update spreadsheet with new data                
        update_speakers(speakers_xlsx, data)
